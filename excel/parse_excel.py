#!/usr/bin env python
# coding=utf-8

import openpyxl

# myDict = dict(myTuples)


# print(json.dumps(myDict,ensure_ascii=False))

# for d in c:
#     if d not in myTuples:
#         print(d)

def load_data_from_excel(path):

    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    query = []

    for item in list(sheet.rows)[1:]:
        query.append(item[0].value)

    wb.close()
    return query

if __name__ == "__main__":
    # 创建一个工作簿
    # wb = openpyxl.Workbook()
    # # 创建一个test_case的sheet表单
    # wb.create_sheet('test_case')
    # # 保存为一个xlsx格式的文件
    # wb.save('cases.xlsx')
    # 读取excel中的数据
    # 第一步：打开工作簿
    wb = openpyxl.load_workbook('../data/sellers.xlsx')
    # 第二步：选取表单
    sh = wb['sheet1']
    # 第三步：读取数据
    # 参数 row:行  column：列
    # ce = sh.cell(row = 1,column = 1)   # 读取第一行，第一列的数据
    # print(ce.value)
    # 按行读取数据 list(sh.rows)
    #print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据

    #c = [92592768,412129187,379424083,3287104402,1744921231,2262055331,2830036684,187918344,1787605898,2894454233,1635784664,661559176,451024527,2176017828,60628840,1940724523,57300080,1035757927,360260087,80336826,1870451655,2831879325,108837944,172472316,653243179,520557274,2360209412,1600687454,76547588,10441568,36360759,5730140,393287647,157347552,11029623,77235267,74990102,163563623,34139303,208910091,77502230,278347875,165271463,100000090,55660809,297976734,92934705,92934705,13322765,9429883,311538314,270725242,5729990,313343723,26589639,34974050,220069524481,42072246,261697088,394351867,161290814,337504533,21710130,64172591,16058627,113671790,84990595,191704707,19100812,315735441,50032799,69846160,161536823,41567036,42056755,174966070,5850194,359979914,104965366,11071207,260857994,171097717,384131957,370396379,187919478,398185442,267073579,165286405,252825489,73163328,290271177,206489282,246880913,221567955,297396681,297910786,356458807,342646008,238561466,404586519,27505373,404467142,211374192,52829628,222644384,37247718,257595676,316471124,68509264,220053967654,7172020,357050331,220742919063,220648597265,220642348223,316581599,220147842291,176856273,339470067,227064272,292370271,220666042897,331531632,317072914,177052898,257699114,92511146,11169344,6225995,225577560,224830467,319407065,182123668,53349749,220060679700,230136741,220638627611,288468550,321747318,91726476,259770572,62818971,3661684,198340785,269104424,369792555,220075883869,220193052568,420751774,325829219,46833387,213711379,82169037,282084245,39431672,66667370,74594915,41371885,85870452,173973174,294497632,216561600,220690656548,6355227,6097442,282845002,279827810,29300995,283008719,352721249,400461477,47706531,74375013,264070344,6873882,70261147,164589803,220079219649,337517097,225970232,112211000,397193880,220729134257,344069352,334497090,220657494238,220739553594,286786827,281803772,308341386,268799907,165373404,39600803,362407932,399049705,220325698,35606033,420786975,44633850,70565789,209679012,172067765,11068734,12996088,31410295,31410295,89293840,224722529,10697822,17375228,67486787,196740027,4401054,44162245,257890098,9204273,47882998,289581069,309805695,243189113,35620554,34094661,5935753,296541976,75864401,92577208,61967311,292912645,207708084,206707751,398277859,278627807,92841713,74686699,194655496,157981427,74868112,92862263,28549219,209525467,57580436,19100812,362659687,198689934,37983358,68118585,166505197,34438976,258746198,294859119,208194819,26845188,16622666,8426989,6816827,211351705,171894864,237457940,76194223,50918451,35302638,77615103,388157356,67355894,308135048,68556349,29924468,72398922,163677840,243684805,281066993,38933551,407491449,74298295,88307294,43049040,278589020,339253670,288007104,38645398,79741932,278576972,38791105,26931523,6098509]

    dic={}
    seller_ids = []
    for cases in list(sh.rows)[1:]:
        #str.strip(str(cases[1].value))
        seller_ids.append(cases[1].value)
        # scene_key =  cases[0].value
        # key = cases[5].value
        # value = cases[6].value
        # dic[key]=value

        # print(value)
        # case_data = cases[2].value

        # myTuples.append(tuple([scene_key, str(value)]))
        # myTuples.append(int(value))
        # if value == '#N/A':
        #     pass
        # else:
        #    tmp = value
        #    myTuples.append(tmp)
        # print(scene_key,value)
    print(seller_ids)
    # 关闭工作薄

    # for idx,value in enumerate(myTuples):
    #   sh.cell(row=idx+2, column=3, value=value)

    # wb.save('234.xlsx')

    # idx=0
    # for cases in list(sh.rows)[1:]:
    #     if idx <= 998:
    #         key = str(cases[1].value)
    #         value = dic[key]
    #         print(value)
    #         sh.cell(row=idx + 2, column=3, value=value)
    #         idx = idx+1
    # wb.save('1130.xlsx')
    wb.close()